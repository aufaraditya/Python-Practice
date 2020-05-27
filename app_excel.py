import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename) #Load Excel File
    sheet = wb['Sheet1'] #Access Sheet1 in Excel
    cell = sheet['a1'] #Access Column A, Row 1 in Sheet1
    cell = sheet.cell(1,1) #Same as the above but using cordinate of cell

    for row in range(2, sheet.max_row + 1): #Loop through rows, Start at 2 for excluding the headers
        cell = sheet.cell(row,3) #Get 3rd Column each row
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
               min_row=2, #Start from row 2
               max_row=sheet.max_row, #End to Max row that has value
               min_col=4, #Min + Max Column 4 = Only Column 4
               max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2') #e2 means top left of chart will be on Column E Row 2


    wb.save('transactions_ver2.xlsx')

